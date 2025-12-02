import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pyodbc
from datetime import datetime
import threading
import openpyxl
import pandas as pd

def get_connection():
    try:
        conn = pyodbc.connect(
            "DRIVER={SQL Server};"
            "SERVER=MT0505\SQLEXPRESS;"
            "DATABASE=QLBanHang;"
            "Trusted_Connection=yes;"
        )
        return conn
    except pyodbc.Error as ex:
        sqlstate = ex.args[0]
        messagebox.showerror("Lỗi Kết Nối CSDL", f"Không thể kết nối đến SQL Server:\n{sqlstate}")
        return None

connect_db = get_connection

san_pham_data = {}
khach_hang_data = {}

def load_all_data():
    global san_pham_data, khach_hang_data
    san_pham_data = {} 
    khach_hang_data = {None: "Khách lẻ (Không lưu)"}
    
    for item in tree_sp.get_children():
        tree_sp.delete(item)

    conn = connect_db()
    if conn is None: return
    cur = conn.cursor()
    
    try:
        cur.execute("SELECT masp, tensp, giasp, soluongton FROM SanPham ORDER BY masp") 
        rows_sp = cur.fetchall()
        for row in rows_sp:
            masp, tensp, giasp, soluongton = row
            san_pham_data[masp] = (tensp, giasp, soluongton)
            display_row = (masp, tensp, f"{giasp:,.0f}", soluongton) 
            tree_sp.insert("", tk.END, values=display_row)
            
        combo_sp['values'] = [f"{masp} - {data[0]}" for masp, data in san_pham_data.items()]
        
        cur.execute("SELECT makh, hoten, sdt FROM KhachHang ORDER BY makh") 
        rows_kh = cur.fetchall()
        for row in rows_kh:
            makh, hoten, sdt = row
            khach_hang_data[makh] = (hoten,sdt)
        
        combo_kh['values'] = [khach_hang_data[None]] + [f"{makh} - {data[0]}" for makh, data in khach_hang_data.items() if makh is not None]

    except Exception as e:
        messagebox.showerror("Lỗi Tải Dữ Liệu", str(e))
    finally:
        conn.close()

def clear_input_sp():
    entry_ma_sp.config(state='normal') 
    entry_ma_sp.delete(0, tk.END)
    entry_ten_sp.delete(0, tk.END)
    entry_gia_sp.delete(0, tk.END)
    entry_soluongton.delete(0, tk.END)
    entry_ma_sp.config(state='disabled') 


def them_sp():
    
    tensp = entry_ten_sp.get().strip()
    giasp = entry_gia_sp.get().strip()
    soluongton = entry_soluongton.get().strip()
    
    if not tensp or not giasp or not soluongton:
        messagebox.showwarning("Thiếu Dữ Liệu", "Vui lòng nhập đầy đủ Tên, Giá và Số lượng tồn.")
        return
    
    try:
        giasp = float(giasp)
        soluongton = int(soluongton)
    except ValueError:
        messagebox.showerror("Lỗi Nhập Liệu", "Giá bán và Số lượng tồn phải là số.")
        return

    conn = connect_db()
    if conn is None: return
    cur = conn.cursor()

    try:
        sql = "INSERT INTO SanPham (tensp, giasp, soluongton) VALUES (?, ?, ?)"
        cur.execute(sql, (tensp, giasp, soluongton))
        conn.commit()
        
        messagebox.showinfo("Thành công", "Đã thêm Sản phẩm mới.")
        load_all_data()
        clear_input_sp()
    except pyodbc.Error as err: 
        messagebox.showerror("Lỗi Thêm", f"Lỗi SQL Server: {err}")
    except Exception as e:
        messagebox.showerror("Lỗi Thêm", str(e))
    finally:
        conn.close()

def sua_sp_select(event):
    selected_item = tree_sp.focus()
    if not selected_item: return
    
    values = tree_sp.item(selected_item, 'values')
    if not values: return

    clear_input_sp()
    
    raw_giasp = values[2].replace(',', '') 
    
    entry_ma_sp.config(state='normal') 
    entry_ma_sp.insert(0, values[0]) 
    entry_ten_sp.insert(0, values[1]) 
    entry_gia_sp.insert(0, raw_giasp) 
    entry_soluongton.insert(0, values[3]) 
    
    entry_ma_sp.config(state='disabled') 

def luu_sp():
    entry_ma_sp.config(state='normal') 
    masp = entry_ma_sp.get() 
    entry_ma_sp.config(state='disabled') 
    
    tensp = entry_ten_sp.get().strip()
    giasp = entry_gia_sp.get().strip()
    soluongton = entry_soluongton.get().strip()
    
    if not masp:
        messagebox.showwarning("Lỗi", "Vui lòng chọn Sản phẩm cần cập nhật.")
        return

    if not tensp or not giasp or not soluongton:
        messagebox.showwarning("Thiếu Dữ Liệu", "Vui lòng nhập đầy đủ thông tin.")
        return
        
    try:
        giasp = float(giasp)
        soluongton = int(soluongton)
    except ValueError:
        messagebox.showerror("Lỗi Nhập Liệu", "Giá bán và Số lượng tồn phải là số.")
        return
        
    conn = connect_db()
    if conn is None: return
    cur = conn.cursor()

    try:
        sql = """
            UPDATE SanPham 
            SET tensp=?, giasp=?, soluongton=?
            WHERE masp=?
        """
        cur.execute(sql, (tensp, giasp, soluongton, masp))
        conn.commit()
        messagebox.showinfo("Thành công", f"Đã cập nhật thông tin Sản phẩm có Mã: {masp}.")
        load_all_data()
        clear_input_sp()
    except Exception as e:
        messagebox.showerror("Lỗi Cập Nhật", str(e))
    finally:
        conn.close()

def xoa_sp():
    selected_item = tree_sp.focus()
    if not selected_item:
        messagebox.showwarning("Chưa chọn", "Hãy chọn một Sản phẩm cần xóa.")
        return
    
    masp = tree_sp.item(selected_item, 'values')[0] 

    confirm = messagebox.askyesno("Xác nhận", f"Bạn có chắc muốn xóa Sản phẩm có Mã: {masp}?")
    if not confirm:
        return

    conn = connect_db()
    if conn is None: return
    cur = conn.cursor()
    
    try:
        sql = "DELETE FROM SanPham WHERE masp=?"
        cur.execute(sql, (masp,))
        conn.commit()
        messagebox.showinfo("Thành công", "Đã xóa Sản phẩm.")
        load_all_data()
        clear_input_sp()
    except Exception as e:
        messagebox.showerror("Lỗi Xóa", f"Lỗi: {str(e)}\n(Kiểm tra xem Sản phẩm có đang tồn tại trong các Đơn hàng không.)") 
    finally:
        conn.close()

def them_khach_hang():
    tenkh = entry_ten_kh.get().strip()
    sdt = entry_sdt_kh.get().strip()
    
    if not tenkh or not sdt:
        messagebox.showwarning("Thiếu Dữ Liệu", "Vui lòng nhập đầy đủ Tên và SĐT Khách hàng.")
        return
        
    conn = connect_db()
    if conn is None: return
    cur = conn.cursor()

    try:
        cur.execute("SELECT makh FROM KhachHang WHERE sdt=?", (sdt,))
        if cur.fetchone():
            messagebox.showwarning("Trùng Lặp", "Số điện thoại này đã tồn tại.")
            return

        sql = "INSERT INTO KhachHang (hoten, sdt) VALUES (?, ?); SELECT SCOPE_IDENTITY()"
        cur.execute(sql, (hoten, sdt))
        
        new_makh_raw = cur.fetchone()
        new_makh = int(new_makh_raw[0])
        conn.commit()
        
        messagebox.showinfo("Thành công", f"Đã thêm Khách hàng mới (Mã: {new_makh}).")
        
        load_all_data() 
        combo_kh.set(f"{new_makh} - {hoten}")
        entry_ten_kh.delete(0, tk.END)
        entry_sdt_kh.delete(0, tk.END)
        
    except pyodbc.Error as err: 
        messagebox.showerror("Lỗi Thêm", f"Lỗi SQL Server: {err}")
    except Exception as e:
        messagebox.showerror("Lỗi Thêm", str(e))
    finally:
        conn.close()

cthd_temp = []

def update_cthd_tree():
    for item in tree_cthd.get_children():
        tree_cthd.delete(item)
    
    tong_tien = 0
    for i, item in enumerate(cthd_temp):
        masp, tensp, giasp, soluong = item
        thanh_tien = float(giasp) * int(soluong)
        tong_tien += thanh_tien
        tree_cthd.insert("", tk.END, values=(masp, tensp, f"{float(giasp):,.0f}", soluong, f"{thanh_tien:,.0f}"))

    lbl_tong_tien.config(text=f"TỔNG THANH TOÁN: {tong_tien:,.0f} VND")
    
def xuatExecel(madh, cthd_data, tong_tien):
    
    ten_file_mac_dinh = f"HoaDon_{madh}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        initialfile=ten_file_mac_dinh,
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    
    if not file_path:
        return False 
    
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"HD #{madh}"

        ws['A1'] = "CỬA HÀNG BÁN LẺ XYZ"
        ws['A2'] = "HOÁ ĐƠN BÁN HÀNG"
        ws['A3'] = f"Mã Đơn Hàng: {madh}"
        ws['A4'] = f"Ngày: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        
        headers = ["STT", "Mã SP", "Tên Sản phẩm", "Đơn Giá", "Số lượng", "Thành Tiền"]
        ws.append([])
        ws.append(headers)
        
        row_num = 0
        for i, item in enumerate(cthd_data):
            masp, tensp, giasp, soluong = item
            thanh_tien = float(giasp) * int(soluong)
            
            display_row = [
                i + 1,
                masp, 
                tensp, 
                float(giasp), 
                soluong, 
                thanh_tien
            ]
            ws.append(display_row)
            row_num = ws.max_row
        
        ws.cell(row=row_num + 2, column=5, value="TỔNG CỘNG:")
        ws.cell(row=row_num + 2, column=6, value=tong_tien).number_format = '#,##0'
        
        for row in ws.iter_rows(min_row=7, max_row=ws.max_row):
            row[3].number_format = '#,##0' 
            row[5].number_format = '#,##0'

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
            
        wb.save(file_path)
        messagebox.showinfo("Xuất Excel Thành Công", f"Đã xuất hóa đơn #{madh} ra file:\n{file_path}")
        return True
    
    except Exception as e:
        messagebox.showerror("Lỗi Xuất Excel", f"Không thể xuất file Excel:\n{str(e)}")
        return False
      
def them_sp_vao_don():
    
    selected_sp_str = combo_sp.get()
    if not selected_sp_str:
        messagebox.showwarning("Lỗi", "Vui lòng chọn Sản phẩm.")
        return
    
    try:
        masp = int(selected_sp_str.split(' - ')[0])
        soluong = int(entry_soluong_ban.get())
        if soluong <= 0:
            messagebox.showwarning("Lỗi", "Số lượng phải lớn hơn 0.")
            return
    except ValueError:
        messagebox.showerror("Lỗi", "Số lượng không hợp lệ.")
        return

    if masp not in san_pham_data:
        messagebox.showerror("Lỗi", "Sản phẩm không tồn tại.")
        return

    tensp, giasp, soluongton = san_pham_data[masp]
    
    new_soluong = soluong
    found = False
    for i, item in enumerate(cthd_temp):
        if item[0] == masp: 
            new_soluong = item[3] + soluong
            found = True
            break
            
    if new_soluong > soluongton:
        messagebox.showwarning("Lỗi Tồn Kho", f"Số lượng tối đa có thể bán là {soluongton}.")
        return

    if found:
        cthd_temp[i] = (masp, tensp, giasp, new_soluong)
    else:
        cthd_temp.append((masp, tensp, giasp, soluong))

    update_cthd_tree()
    entry_soluong_ban.delete(0, tk.END)
    entry_soluong_ban.insert(0, "1")

def lap_don_hang():
    if not cthd_temp:
        messagebox.showwarning("Lỗi", "Đơn hàng rỗng.")
        return
        
    tong_tien = sum(float(item[2]) * int(item[3]) for item in cthd_temp)

    selected_kh_str = combo_kh.get()
    makh_val = None
    if selected_kh_str and "Khách lẻ" not in selected_kh_str:
        try:
            makh_val = int(selected_kh_str.split(' - ')[0])
        except ValueError:
            makh_val = None
            
    conn = connect_db()
    if conn is None: return
    cur = conn.cursor()
    
    try:
        ngaydat = datetime.now().strftime('%Y-%m-%d')
        sql_dh = "INSERT INTO DonHang (ngaydat, tongtien, makh) VALUES (?, ?, ?); SELECT SCOPE_IDENTITY()"
        cur.execute(sql_dh, (ngaydat, tong_tien, makh_val)) 
        
        if cur.nextset():
            madh_raw = cur.fetchone()
        else:
            raise Exception("Không thể chuyển đến tập kết quả SELECT SCOPE_IDENTITY().")

        madh = int(madh_raw[0]) if madh_raw is not None and madh_raw[0] is not None else None

        if madh is None:
            raise Exception("Lỗi: Mã Đơn Hàng mới là NULL sau khi chèn.")
        
        for masp, _, giasp, soluong in cthd_temp:
            sql_cthd = "INSERT INTO CTHD (madh, masp, soluong, dongia) VALUES (?, ?, ?, ?)"
            cur.execute(sql_cthd, (madh, masp, soluong, giasp))
            
            sql_update_sp = "UPDATE SanPham SET soluongton = soluongton - ? WHERE masp = ?"
            cur.execute(sql_update_sp, (soluong, masp))
            
        conn.commit() 
        messagebox.showinfo("Thành công", f"Đã lập thành công Đơn hàng #{madh}. Tổng tiền: {tong_tien:,.0f} VND")
        data_to_export = list(cthd_temp)
        xuatExecel(madh, data_to_export ,tong_tien)

        clear_don_hang()
        load_all_data() 

    except Exception as e:
        if conn:
            conn.rollback() 
        messagebox.showerror("Lỗi Lập Đơn Hàng", str(e))
    finally:
        conn.close()
        
def xoa_sp_trong_don():
    selected_item = tree_cthd.focus()
    if not selected_item:
        messagebox.showwarning("Chưa chọn", "Hãy chọn một mục cần xóa khỏi đơn hàng.")
        return
    
    try:
        masp_can_xoa = int(tree_cthd.item(selected_item, 'values')[0])
    except:
        messagebox.showerror("Lỗi", "Không thể lấy mã sản phẩm.")
        return
    
    global cthd_temp
    cthd_temp = [item for item in cthd_temp if item[0] != masp_can_xoa]
    
    update_cthd_tree()
    
def clear_don_hang():
    global cthd_temp
    cthd_temp = []
    update_cthd_tree()
    entry_soluong_ban.delete(0, tk.END)
    entry_soluong_ban.insert(0, "1")
    combo_kh.set("Khách lẻ ")
    
def reset_sp_db_task():
    conn = connect_db()
    if conn is None: 
        return
    cur = conn.cursor()

    try:
        cur.execute("BEGIN TRANSACTION")
        
        cur.execute("DELETE FROM CTHD")
        
        cur.execute("DELETE FROM SanPham")
        
        cur.execute("DBCC CHECKIDENT ('SanPham', RESEED, 0)")
        
        conn.commit()
        
        root.after(0, lambda: messagebox.showinfo("Thành công", "Đã xóa toàn bộ Sản phẩm và reset Mã SP về 1."))
        root.after(0, load_all_data)
        root.after(0, clear_input_sp)
        
    except pyodbc.Error as err: 
        conn.rollback()
        root.after(0, lambda: messagebox.showerror("Lỗi CSDL", f"Không thể reset bảng Sản phẩm. Lỗi SQL: {err}"))
    except Exception as e:
        conn.rollback()
        root.after(0, lambda: messagebox.showerror("Lỗi", str(e)))
    finally:
        conn.close()

def reset_sp_table():
    
    confirm = messagebox.askyesno(
        "XÁC NHẬN RESET", 
        "CẢNH BÁO: Thao tác này sẽ XÓA TOÀN BỘ dữ liệu Sản phẩm và RESET Mã SP về 1.\n\nBạn có chắc chắn muốn tiếp tục không?"
    )
    if not confirm:
        return

    reset_thread = threading.Thread(target=reset_sp_db_task)
    reset_thread.start()    

root = tk.Tk()
root.title("QUẢN LÝ BÁN HÀNG")
root.geometry("1100x650") 
root.resizable(False, False)
main_frame = tk.Frame(root)
main_frame.pack(padx=10, pady=10, fill="both", expand=True)

frame_sp = tk.Frame(main_frame, bd=2, relief=tk.GROOVE)
frame_sp.pack(side=tk.LEFT, padx=10, fill="y", expand=False)

tk.Label(frame_sp, text="QUẢN LÝ SẢN PHẨM", font=("Arial", 14, "bold"), fg="navy").pack(pady=10)

frame_input_sp = tk.Frame(frame_sp)
frame_input_sp.pack(pady=5, padx=5, fill="x")

tk.Label(frame_input_sp, text="Mã SP:", width=10).grid(row=0, column=0, padx=5, pady=2, sticky="w")
entry_ma_sp = tk.Entry(frame_input_sp, width=20)
entry_ma_sp.grid(row=0, column=1, padx=5, pady=2, sticky="w")
entry_ma_sp.config(state='disabled')

tk.Label(frame_input_sp, text="Tên SP:", width=10).grid(row=1, column=0, padx=5, pady=2, sticky="w")
entry_ten_sp = tk.Entry(frame_input_sp, width=20)
entry_ten_sp.grid(row=1, column=1, padx=5, pady=2, sticky="w")

tk.Label(frame_input_sp, text="Giá Bán:", width=10).grid(row=2, column=0, padx=5, pady=2, sticky="w")
entry_gia_sp = tk.Entry(frame_input_sp, width=20)
entry_gia_sp.grid(row=2, column=1, padx=5, pady=2, sticky="w")

tk.Label(frame_input_sp, text="Tồn Kho:", width=10).grid(row=3, column=0, padx=5, pady=2, sticky="w")
entry_soluongton = tk.Entry(frame_input_sp, width=20)
entry_soluongton.grid(row=3, column=1, padx=5, pady=2, sticky="w")

frame_btn_sp = tk.Frame(frame_sp)
frame_btn_sp.pack(pady=10, padx=5)

tk.Button(frame_btn_sp, text="Thêm", width=10, command=them_sp).grid(row=0, column=0, padx=5)
tk.Button(frame_btn_sp, text="Lưu/Cập nhật", width=12, command=luu_sp).grid(row=0, column=1, padx=5)
tk.Button(frame_btn_sp, text="Làm mới", width=10, command=clear_input_sp).grid(row=0, column=2, padx=5)
tk.Button(frame_btn_sp, text="Xóa", width=10, command=xoa_sp).grid(row=0, column=3, padx=5)

lbl_list_sp = tk.Label(frame_sp, text="Danh sách Sản phẩm", font=("Arial", 10, "bold"))
lbl_list_sp.pack(pady=(5, 2))

tree_frame_sp = tk.Frame(frame_sp)
tree_frame_sp.pack(padx=5, pady=5, fill='both', expand=True)

tree_scroll_sp = ttk.Scrollbar(tree_frame_sp, orient="vertical")
tree_scroll_sp.pack(side="right", fill="y")

tree_sp = ttk.Treeview(tree_frame_sp, columns=("masp", "tensp", "giasp", "soluongton"),show="headings", yscrollcommand=tree_scroll_sp.set)
tree_scroll_sp.config(command=tree_sp.yview)

tree_sp.heading("masp", text="Mã SP")
tree_sp.heading("tensp", text="Tên Sản phẩm")
tree_sp.heading("giasp", text="Giá Bán")
tree_sp.heading("soluongton", text="Tồn Kho")

tree_sp.column("masp", width=60, anchor=tk.CENTER)
tree_sp.column("tensp", width=150)
tree_sp.column("giasp", width=90, anchor=tk.E)
tree_sp.column("soluongton", width=70, anchor=tk.CENTER)

tree_sp.pack(fill="both", expand=True)
tree_sp.bind('<<TreeviewSelect>>', sua_sp_select)

frame_dh = tk.Frame(main_frame, bd=2, relief=tk.GROOVE)
frame_dh.pack(side=tk.LEFT, padx=10, fill="both", expand=True)

tk.Label(frame_dh, text="BÁN HÀNG TẠI QUẦY", font=("Arial", 14, "bold"), fg="darkgreen").pack(pady=10)

frame_add_sp = tk.Frame(frame_dh)
frame_add_sp.pack(pady=5, padx=10, fill="x")

tk.Label(frame_add_sp, text="Chọn SP:", width=10).grid(row=0, column=0, padx=5, pady=5, sticky="w")
combo_sp = ttk.Combobox(frame_add_sp, width=35, state="readonly")
combo_sp.grid(row=0, column=1, padx=5, pady=5, sticky="w")

tk.Label(frame_add_sp, text="SL:", width=5).grid(row=0, column=2, padx=5, pady=5, sticky="w")
entry_soluong_ban = tk.Entry(frame_add_sp, width=5)
entry_soluong_ban.insert(0, "1")
entry_soluong_ban.grid(row=0, column=3, padx=5, pady=5, sticky="w")

tk.Button(frame_add_sp, text="➕ Thêm", command=them_sp_vao_don, width=10).grid(row=0, column=4, padx=10, pady=5)

def on_sp_select(event):
    selected_sp_str = combo_sp.get()
    if not selected_sp_str: return
    try:
        masp = int(selected_sp_str.split(' - ')[0]) 
        if masp in san_pham_data:
            tensp, giasp, soluongton = san_pham_data[masp]
            lbl_sp_info.config(text=f"Giá: {float(giasp):,.0f} VND | Tồn: {soluongton}")
        else:
            lbl_sp_info.config(text="Giá: N/A | Tồn: N/A")
    except:
        lbl_sp_info.config(text="Giá: N/A | Tồn: N/A")


combo_sp.bind('<<ComboboxSelected>>', on_sp_select)
lbl_sp_info = tk.Label(frame_add_sp, text="Giá: N/A | Tồn: N/A", fg="gray")
lbl_sp_info.grid(row=1, column=1, padx=5, sticky="w", columnspan=3)

frame_kh = tk.Frame(frame_dh, bd=1, relief=tk.SOLID, padx=5, pady=5)
frame_kh.pack(pady=10, padx=10, fill="x")
tk.Label(frame_kh, text="THÔNG TIN KHÁCH HÀNG", font=("Arial", 10, "bold")).pack(pady=5)

frame_kh_input = tk.Frame(frame_kh)
frame_kh_input.pack()

tk.Label(frame_kh_input, text="Khách hàng hiện tại:", width=18).grid(row=0, column=0, padx=5, pady=2, sticky="w")
combo_kh = ttk.Combobox(frame_kh_input, width=30, state="readonly")
combo_kh.grid(row=0, column=1, padx=5, pady=2, sticky="w")
combo_kh.set("Khách lẻ (Không lưu)")

tk.Label(frame_kh_input, text="Tên KH Mới:", width=18).grid(row=1, column=0, padx=5, pady=2, sticky="w")
entry_ten_kh = tk.Entry(frame_kh_input, width=32)
entry_ten_kh.grid(row=1, column=1, padx=5, pady=2, sticky="w")

tk.Label(frame_kh_input, text="SĐT Mới:", width=18).grid(row=2, column=0, padx=5, pady=2, sticky="w")
entry_sdt_kh = tk.Entry(frame_kh_input, width=32)
entry_sdt_kh.grid(row=2, column=1, padx=5, pady=2, sticky="w")

tk.Button(frame_kh_input, text="Lưu KH Mới", command=them_khach_hang, width=15).grid(row=1, column=2, rowspan=2, padx=10)

lbl_list_cthd = tk.Label(frame_dh, text="Chi tiết Đơn hàng", font=("Arial", 10, "bold"))
lbl_list_cthd.pack(pady=(5, 2))

tree_frame_cthd = tk.Frame(frame_dh)
tree_frame_cthd.pack(padx=10, pady=5, fill='both', expand=True)

tree_scroll_cthd = ttk.Scrollbar(tree_frame_cthd, orient="vertical")
tree_scroll_cthd.pack(side="right", fill="y")

tree_cthd = ttk.Treeview(tree_frame_cthd, columns=("masp", "tensp", "giasp", "soluong", "thanhtien"), 
                             show="headings", yscrollcommand=tree_scroll_cthd.set)
tree_scroll_cthd.config(command=tree_cthd.yview)

tree_cthd.heading("masp", text="Mã SP")
tree_cthd.heading("tensp", text="Tên Sản phẩm")
tree_cthd.heading("giasp", text="Đơn Giá")
tree_cthd.heading("soluong", text="Số lượng")
tree_cthd.heading("thanhtien", text="Thành tiền")

tree_cthd.column("masp", width=60, anchor=tk.CENTER)
tree_cthd.column("tensp", width=200)
tree_cthd.column("giasp", width=100, anchor=tk.E)
tree_cthd.column("soluong", width=70, anchor=tk.CENTER)
tree_cthd.column("thanhtien", width=120, anchor=tk.E)

tree_cthd.pack(fill="both", expand=True)

frame_footer_dh = tk.Frame(frame_dh)
frame_footer_dh.pack(pady=10, padx=10, fill="x")
lbl_tong_tien = tk.Label(frame_footer_dh, text="TỔNG THANH TOÁN: 0 VND", font=("Arial", 14, "bold"), fg="red")
lbl_tong_tien.pack(side=tk.LEFT, padx=10)
tk.Button(frame_footer_dh, text=" Xóa SP", width=10, command=xoa_sp_trong_don).pack(side=tk.RIGHT, padx=5)
tk.Button(frame_footer_dh, text=" Làm mới", width=10, command=clear_don_hang).pack(side=tk.RIGHT, padx=5)
tk.Button(frame_footer_dh, text=" LẬP ĐƠN", width=15, bg="green", fg="white", 
          font=("Arial", 12, "bold"), command=lap_don_hang).pack(side=tk.RIGHT, padx=15)

load_all_data() 
root.mainloop()